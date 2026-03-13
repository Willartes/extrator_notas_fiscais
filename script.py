"""
Extrator de Notas Fiscais de PDF para Excel
Relatório: Valores Mensais de Retenções por Participante
SAUDE SANTA TEREZA LTDA

Particularidade do PDF: o pdfplumber cola a chave numérica diretamente
com a data na mesma linha, sem espaço.
Exemplo de linha extraída:
  '2216901/02/2026 NFSE 59339 400,00 5952 2,60 5952 12,00 5952 4,00 0,00 0,00'

Ordem real das colunas após o número da nota (confirmada pelo diagnóstico):
  Valor_Contabil | cod | PIS | cod | Cofins | cod | CSLL | [cod] | IRRF | INSS
"""

import sys
import re
import os

# ============================================================
# CONFIGURAÇÃO PRINCIPAL
# ============================================================
CAMINHO_PDF   = "notas.pdf"
CAMINHO_EXCEL = "notas_fiscais.xlsx"
# ============================================================


def verificar_dependencias():
    faltando = []
    for pacote in ["pdfplumber", "openpyxl"]:
        try:
            __import__(pacote)
        except ImportError:
            faltando.append(pacote)
    if faltando:
        print(f"[ERRO] Pacotes não encontrados: {', '.join(faltando)}")
        print(f"[DICA] Execute: pip install {' '.join(faltando)}")
        sys.exit(1)


verificar_dependencias()

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# PADRÕES REGEX
# ============================================================

# CORREÇÃO PRINCIPAL: a linha começa com dígitos da chave colados na data.
# Exemplos reais:
#   '2216901/02/2026 NFSE 59339 400,00 5952 2,60 5952 12,00 5952 4,00 0,00 0,00'
#   '2218401/02/2026 NFSE 59101 6.500,00 5952 42,25 5952 195,00 5952 65,00 1708 97,50 0,00'
#
# Padrão: zero ou mais dígitos (chave), seguidos de DD/MM/AAAA, espaço, NFSE, espaço, número
RE_LINHA_NOTA = re.compile(
    r"^\d*"                        # chave colada (0 ou mais dígitos)
    r"(\d{2}/\d{2}/\d{4})"        # data DD/MM/AAAA
    r"\s+NFSE\s+"                  # espécie
    r"(\d+)"                       # número da nota
    r"\s+(.+)$"                    # resto da linha com os valores
)

# Valor monetário brasileiro: 0,00 | 400,00 | 6.500,00 | 1.281,75
RE_VALOR = re.compile(r"^\d{1,3}(?:\.\d{3})*,\d{2}$")

# Código de receita: exatamente 4 dígitos
RE_COD_RECEITA = re.compile(r"^\d{4}$")

# Cabeçalho de cliente
RE_CLIENTE = re.compile(r"^Cliente:\s*\d+\s*-\s*(.+?)(?:\s{2,}|\s+CNPJ:)", re.IGNORECASE)
RE_CLIENTE_FB = re.compile(r"^Cliente:\s*\d+\s*-\s*(.+)", re.IGNORECASE)

# Cidade na linha de endereço
RE_CIDADE = re.compile(r"Cidade:\s*(.+)", re.IGNORECASE)

# Linha de totais (ignorar)
RE_TOTAIS = re.compile(r"^\s*Totais\b", re.IGNORECASE)

# Linha de cabeçalho de coluna (ignorar)
RE_CABECALHO_COL = re.compile(
    r"ChaveData|Espécie|Série|Número|receita|Códigoreceita|Valor contábil",
    re.IGNORECASE
)


# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def limpar_valor(texto):
    """
    Converte string monetária brasileira para float.
    '6.500,00' -> 6500.0 | '0,00' -> 0.0
    """
    texto = str(texto).strip()
    if not texto:
        return 0.0
    try:
        return float(texto.replace(".", "").replace(",", "."))
    except ValueError:
        return 0.0


def extrair_valores(trecho):
    """
    Recebe o trecho de texto após o número da nota e retorna
    apenas os tokens que são valores monetários (descarta códigos
    de receita de 4 dígitos e qualquer outro token).

    Exemplo de trecho:
      '400,00 5952 2,60 5952 12,00 5952 4,00 0,00 0,00'
    Retorna:
      ['400,00', '2,60', '12,00', '4,00', '0,00', '0,00']

    Mapeamento das colunas (ordem confirmada pelo diagnóstico):
      [0] Valor_Contabil
      [1] PIS
      [2] Cofins
      [3] CSLL
      [4] IRRF
      [5] INSS
    """
    tokens = trecho.split()
    return [t for t in tokens if RE_VALOR.match(t)]


def parsear_linha_nota(linha, razao_social, cidade):
    """
    Faz o parse completo de uma linha de nota fiscal.

    A regex RE_LINHA_NOTA extrai:
      group(1) = data (não usada na saída)
      group(2) = número da nota
      group(3) = trecho com valores e códigos de receita

    Mapeamento final dos valores filtrados:
      [0] Valor_Contabil  <- primeira coluna de valor
      [1] PIS
      [2] Cofins
      [3] CSLL
      [4] IRRF            <- 0,00 quando não há retenção (sem código antes)
      [5] INSS            <- sempre 0,00 neste relatório
    """
    m = RE_LINHA_NOTA.match(linha.strip())
    if not m:
        return None

    numero_nota = m.group(2)
    trecho      = m.group(3)

    valores = extrair_valores(trecho)

    def get(idx):
        try:
            return limpar_valor(valores[idx])
        except IndexError:
            return 0.0

    return {
        "Numero_Nota":    numero_nota,
        "Razao_Social":   razao_social,
        "Cidade":         cidade,
        "Valor_Contabil": get(0),
        "PIS":            get(1),
        "Cofins":         get(2),
        "CSLL":           get(3),
        "IRRF":           get(4),
        "INSS":           get(5),
    }


def extrair_razao_social(linha):
    """Extrai razão social da linha de cabeçalho do cliente."""
    m = RE_CLIENTE.search(linha)
    if m:
        return m.group(1).strip()
    m = RE_CLIENTE_FB.search(linha)
    if m:
        nome = re.sub(r"\s+CNPJ:.+", "", m.group(1), flags=re.IGNORECASE)
        return nome.strip()
    return linha.strip()


def extrair_cidade(linha):
    """Extrai cidade da linha de endereço."""
    m = RE_CIDADE.search(linha)
    return m.group(1).strip() if m else ""


def deve_ignorar(linha):
    """
    Retorna True para linhas que não devem ser processadas:
    totais, cabeçalhos de coluna, rodapés e linhas de dígito
    solto (segunda parte da chave).
    """
    if RE_TOTAIS.match(linha):
        return True
    if RE_CABECALHO_COL.search(linha):
        return True
    # Rodapé do sistema
    if "ANALISE EMPRESARIAL" in linha or "SCI Ambiente" in linha:
        return True
    # Linha com apenas dígitos (segundo fragmento da chave, ex: '4' ou '3')
    if re.match(r"^\d+$", linha.strip()):
        return True
    # Cabeçalho do relatório
    if linha.startswith("Valores Mensais") or linha.startswith("Empresa:"):
        return True
    return False


# ============================================================
# EXTRAÇÃO PRINCIPAL
# ============================================================

def extrair_dados_pdf(caminho_pdf):
    """
    Percorre todas as páginas do PDF e extrai os dados de cada nota fiscal.

    Fluxo por linha:
      1. Ignora linhas de ruído (totais, cabeçalhos, rodapés, dígitos soltos)
      2. Detecta cabeçalho de cliente -> atualiza razao_social_atual
      3. Detecta linha de endereço   -> atualiza cidade_atual
      4. Detecta linha de nota       -> parseia e adiciona ao resultado
    """
    registros          = []
    clientes_vistos    = set()
    razao_social_atual = ""
    cidade_atual       = ""
    falhas_parse       = 0

    print(f"\n[INFO] Abrindo: {caminho_pdf}")

    with pdfplumber.open(caminho_pdf) as pdf:
        total_paginas = len(pdf.pages)
        print(f"[INFO] Total de páginas: {total_paginas}\n")

        for num_pag, pagina in enumerate(pdf.pages, start=1):
            texto = pagina.extract_text()
            if not texto:
                print(f"[AVISO] Página {num_pag} sem texto — pulando.")
                continue

            for linha in texto.split("\n"):
                linha_s = linha.strip()
                if not linha_s:
                    continue

                # ── Filtra ruído ──────────────────────────────────────
                if deve_ignorar(linha_s):
                    continue

                # ── Cabeçalho de cliente ──────────────────────────────
                if linha_s.lower().startswith("cliente:"):
                    razao_social_atual = extrair_razao_social(linha_s)
                    clientes_vistos.add(razao_social_atual)
                    continue

                # ── Endereço / cidade ─────────────────────────────────
                if "Cidade:" in linha_s:
                    cidade_atual = extrair_cidade(linha_s)
                    continue

                # ── Linha de nota fiscal ──────────────────────────────
                if RE_LINHA_NOTA.match(linha_s):
                    if not razao_social_atual:
                        print(f"[AVISO] Nota sem cliente associado na pág.{num_pag}: {linha_s[:60]}")
                        continue

                    reg = parsear_linha_nota(linha_s, razao_social_atual, cidade_atual)
                    if reg:
                        registros.append(reg)
                    else:
                        falhas_parse += 1
                        if falhas_parse <= 15:
                            print(f"[AVISO] Parse falhou pág.{num_pag}: {linha_s[:80]}")

    print(f"[RESULTADO] Clientes identificados   : {len(clientes_vistos)}")
    print(f"[RESULTADO] Notas fiscais extraídas  : {len(registros)}")
    if falhas_parse:
        print(f"[RESULTADO] Linhas com falha de parse: {falhas_parse}")

    return registros


# ============================================================
# GERAÇÃO DO EXCEL
# ============================================================

def gerar_excel(registros, caminho_excel):
    """Gera o arquivo Excel com cabeçalho estilizado e dados formatados."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Notas Fiscais"

    COLUNAS = [
        "Numero_Nota",
        "Razao_Social",
        "Cidade",
        "Valor_Contabil",
        "PIS",
        "Cofins",
        "CSLL",
        "IRRF",
        "INSS",
    ]

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")

    # Cabeçalho
    for col_idx, nome in enumerate(COLUNAS, start=1):
        cel = ws.cell(row=1, column=col_idx, value=nome)
        cel.font      = hdr_font
        cel.fill      = hdr_fill
        cel.alignment = hdr_align

    # Dados
    for row_idx, reg in enumerate(registros, start=2):
        ws.cell(row=row_idx, column=1, value=reg["Numero_Nota"])
        ws.cell(row=row_idx, column=2, value=reg["Razao_Social"])
        ws.cell(row=row_idx, column=3, value=reg["Cidade"])
        ws.cell(row=row_idx, column=4, value=reg["Valor_Contabil"])
        ws.cell(row=row_idx, column=5, value=reg["PIS"])
        ws.cell(row=row_idx, column=6, value=reg["Cofins"])
        ws.cell(row=row_idx, column=7, value=reg["CSLL"])
        ws.cell(row=row_idx, column=8, value=reg["IRRF"])
        ws.cell(row=row_idx, column=9, value=reg["INSS"])

    # Formato numérico nas colunas de valor
    fmt = '#,##0.00'
    for row_idx in range(2, len(registros) + 2):
        for col_idx in range(4, 10):
            ws.cell(row=row_idx, column=col_idx).number_format = fmt

    # Largura das colunas
    for col_idx, larg in zip(range(1, 10), [15, 55, 25, 18, 12, 12, 12, 12, 12]):
        letra = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letra].width = larg

    wb.save(caminho_excel)
    print(f"\n[SUCESSO] Arquivo gerado  : {caminho_excel}")
    print(f"[SUCESSO] Linhas de dados : {len(registros)}")


# ============================================================
# DIAGNÓSTICO (mantenha max_paginas=0 em produção)
# ============================================================

def diagnosticar(caminho_pdf, max_paginas=0):
    if max_paginas <= 0:
        return
    print("\n" + "="*60)
    print(f"DIAGNÓSTICO — primeiras {max_paginas} página(s)")
    print("="*60)
    with pdfplumber.open(caminho_pdf) as pdf:
        for i, pag in enumerate(pdf.pages[:max_paginas], 1):
            print(f"\n--- PÁGINA {i} ---")
            texto = pag.extract_text()
            if texto:
                for ln in texto.split("\n"):
                    print(repr(ln))
            else:
                print("  [sem texto]")
    print("="*60 + "\n")


# ============================================================
# PONTO DE ENTRADA
# ============================================================

if __name__ == "__main__":

    if not os.path.isfile(CAMINHO_PDF):
        print(f"[ERRO] PDF não encontrado: {CAMINHO_PDF}")
        sys.exit(1)

    # Mude para max_paginas=2 se quiser inspecionar o texto bruto novamente
    diagnosticar(CAMINHO_PDF, max_paginas=0)

    registros = extrair_dados_pdf(CAMINHO_PDF)

    if not registros:
        print("\n[DICA] Nenhuma nota extraída. Ative o diagnóstico para inspecionar.")
        sys.exit(1)

    gerar_excel(registros, CAMINHO_EXCEL)